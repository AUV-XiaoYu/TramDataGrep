"""
导出模块
支持 Excel (.xlsx) 和 CSV (.csv) 两种格式的流式生成
"""
import csv as _csv
from datetime import datetime, date
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def generate_excel(columns, row_generator):
    """
    流式生成 Excel 文件

    Args:
        columns: 列名列表
        row_generator: 行数据生成器，每次 yield 一批行

    Yields:
        bytes: Excel 文件的字节块（64KB 每块）
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="数据")

    # 写表头
    ws.append(columns)

    # 流式写入数据行（列名已由调用方通过 next(row_generator) 消费）
    for batch in row_generator:

        for row in batch:
            values = [_serialize_value(v) for v in row]
            ws.append(values)

    # 输出到 buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # 分块返回
    while True:
        chunk = buffer.read(65536)  # 64KB
        if not chunk:
            break
        yield chunk


def generate_csv(columns, row_generator):
    """
    流式生成 CSV 文件（UTF-8 with BOM，兼容 Excel 中文环境）

    Args:
        columns: 列名列表
        row_generator: 行数据生成器，每次 yield 一批行

    Yields:
        bytes: CSV 文件的字节块
    """
    # 先写 BOM，让 Excel 正确识别 UTF-8 编码的中文
    yield "﻿".encode("utf-8")

    output = StringIO()
    writer = _csv.writer(output)

    # 写表头
    writer.writerow(columns)

    # 流式写入数据行（列名已由调用方通过 next(row_generator) 消费）
    for batch in row_generator:
        for row in batch:
            writer.writerow([_serialize_value(v) for v in row])
        # 每批刷一次，避免 output 无限增长
        chunk = output.getvalue().encode("utf-8")
        output.truncate(0)
        output.seek(0)
        if chunk:
            yield chunk

    # 尾量
    tail = output.getvalue().encode("utf-8")
    if tail:
        yield tail


def _serialize_value(value):
    """将数据库值转为 Excel 兼容格式"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return value